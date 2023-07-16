import tabula
import pandas as pd
from fpdf import FPDF
import openpyxl
import os

def pdf_to_xlsx(file_name):
    src = os.getcwd() + '\\uploads\\' + file_name
    dest = os.getcwd() + '\\processed_files\\' + file_name[:-3] + 'docx'

    csv = tabula.read_pdf(src, pages='all')

    excel_writer = pd.ExcelWriter(dest)
    for i, df in enumerate(csv):
        df.to_excel(excel_writer, sheet_name='Table_' + str(i), index=False)
    excel_writer.close()

    return csv

def xlsx_to_pdf(file_name):
    src = 'uploads\\' + file_name
    dest = 'processed_files\\' + file_name[:-3] + 'pdf'

    wb = openpyxl.load_workbook(src)
    sheet = wb.active

    pdf = FPDF()
    for row in sheet.iter_rows(values_only=True):
        pdf.cell(40, 10, ' '.join(row))

    pdf.output(dest)


if __name__ == '__main__':
    file_name = 'abc.pdf'

    pdf_to_xlsx(file_name)